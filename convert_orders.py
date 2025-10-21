#!/usr/bin/env python3
"""
Convert Staples order JSON data to TSV and XLSX formats.
Handles both in-store and online order formats.
"""

import json
import csv
import sys
import argparse
from pathlib import Path
from datetime import datetime
from typing import List, Dict, Any

try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter
    XLSX_AVAILABLE = True
except ImportError:
    XLSX_AVAILABLE = False
    print("Warning: openpyxl not installed. Install with: pip install openpyxl")
    print("Only TSV output will be generated.")


def parse_order_lines(order: Dict[str, Any]) -> List[Dict[str, Any]]:
    """Parse order lines from both in-store and online formats."""
    rows = []

    # Common order fields
    order_number = order.get('orderNumber', '')
    order_date = order.get('orderDate', '')
    order_type = order.get('orderType', '')
    order_channel = order.get('orderChannel', order_type)  # Use orderType if channel missing
    customer_number = order.get('customerNumber', '')
    status_desc = order.get('statusDesc', '')
    grand_total = order.get('grandTotal', 0)
    points_issued = order.get('pointsIssued', 0)
    points_redeemed = order.get('pointsRedeemed', 0)

    # Store information (for in-store orders)
    store_number = order.get('storeNumber', '')
    store_address = order.get('storeAddressLine1', '')
    store_city = order.get('storeCity', '')
    store_state = order.get('storeState', '')

    # Process each order line
    for line in order.get('orderLines', []):
        row = {
            'Order Number': order_number,
            'Order Date': order_date,
            'Order Type': order_type,
            'Order Channel': order_channel,
            'Customer Number': customer_number,
            'Order Status': status_desc,
            'Grand Total': grand_total,
            'Points Issued': points_issued,
            'Points Redeemed': points_redeemed,
            'Store Number': store_number,
            'Store Address': store_address,
            'Store City': store_city,
            'Store State': store_state,
            'Product SKU': line.get('productSku', ''),
            'Product Name': line.get('productName', ''),
            'Product Description': line.get('productDesc', ''),
            'Model': line.get('model', ''),
            'Unit Price': line.get('unitPrice', 0),
            'Quantity': line.get('orderQty', 0),
            'Original Quantity': line.get('originalOrderedQty', 0),
            'Line Total': line.get('total', 0),
            'Line Status': line.get('statuses', [{}])[0].get('statusDesc', '') if line.get('statuses') else '',
            'Product URL': line.get('productURL', ''),
            'Out of Stock': line.get('isOutOfStock', False),
            'BOPIS Eligible': line.get('bopisEligible', False),
            'Delivery In Stock': line.get('isDeliveryInstock', False),
        }
        rows.append(row)

    # If no order lines, create a single row with order info
    if not rows:
        rows.append({
            'Order Number': order_number,
            'Order Date': order_date,
            'Order Type': order_type,
            'Order Channel': order_channel,
            'Customer Number': customer_number,
            'Order Status': status_desc,
            'Grand Total': grand_total,
            'Points Issued': points_issued,
            'Points Redeemed': points_redeemed,
            'Store Number': store_number,
            'Store Address': store_address,
            'Store City': store_city,
            'Store State': store_state,
            'Product SKU': '',
            'Product Name': '',
            'Product Description': '',
            'Model': '',
            'Unit Price': 0,
            'Quantity': 0,
            'Original Quantity': 0,
            'Line Total': 0,
            'Line Status': '',
            'Product URL': '',
            'Out of Stock': False,
            'BOPIS Eligible': False,
            'Delivery In Stock': False,
        })

    return rows


def load_json_file(file_path: Path) -> Dict[str, Any]:
    """Load and parse a JSON file."""
    try:
        with open(file_path, 'r', encoding='utf-8') as f:
            return json.load(f)
    except Exception as e:
        print(f"Error loading {file_path}: {e}")
        return {}


def write_tsv(rows: List[Dict[str, Any]], output_file: Path):
    """Write rows to a TSV file."""
    if not rows:
        print("No data to write to TSV")
        return

    # Get all unique keys from all rows
    fieldnames = list(rows[0].keys())

    with open(output_file, 'w', newline='', encoding='utf-8') as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames, delimiter='\t')
        writer.writeheader()
        writer.writerows(rows)

    print(f"✓ Wrote {len(rows)} rows to {output_file}")


def prepare_excel_value(field: str, value: Any):
    """Return converted value along with a flag indicating if it's a datetime."""
    if isinstance(value, datetime):
        return value, True
    if isinstance(value, str) and 'date' in field.lower():
        parsed = parse_order_datetime_value(value)
        if isinstance(parsed, datetime):
            return parsed, True
    return value, False


def write_xlsx(rows: List[Dict[str, Any]], output_file: Path):
    """Write rows to an XLSX file with formatting."""
    if not XLSX_AVAILABLE:
        print("Skipping XLSX generation (openpyxl not installed)")
        return

    if not rows:
        print("No data to write to XLSX")
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "Orders"

    # Header styling
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center")

    # Write headers
    fieldnames = list(rows[0].keys())
    for col_idx, field in enumerate(fieldnames, 1):
        cell = ws.cell(row=1, column=col_idx, value=field)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    # Write data
    for row_idx, row in enumerate(rows, 2):
        for col_idx, field in enumerate(fieldnames, 1):
            raw_value = row.get(field, '')
            value, is_datetime = prepare_excel_value(field, raw_value)
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = Font(size=16)

            if isinstance(value, (int, float)):
                cell.alignment = Alignment(horizontal="right")

            if is_datetime:
                cell.number_format = 'yyyy-mm-dd hh:mm:ss'

    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except Exception:
                pass
        adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
        ws.column_dimensions[column_letter].width = adjusted_width

    # Freeze header row
    ws.freeze_panes = 'A2'

    wb.save(output_file)
    print(f"✓ Wrote {len(rows)} rows to {output_file}")


def write_combined_sheet(all_rows: List[Dict[str, Any]], output_path: Path):
    if not all_rows:
        print("No rows to write in combined workbook.")
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "Orders"

    fieldnames = list(all_rows[0].keys())

    header_fill = PatternFill(start_color="2E74B5", end_color="2E74B5", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=16)
    header_alignment = Alignment(horizontal="center", vertical="center")

    for col_idx, field in enumerate(fieldnames, 1):
        cell = ws.cell(row=1, column=col_idx, value=field)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    for row_idx, row in enumerate(all_rows, 2):
        for col_idx, field in enumerate(fieldnames, 1):
            raw_value = row.get(field, '')
            value, is_datetime = prepare_excel_value(field, raw_value)
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = Font(size=16)
            if is_datetime:
                cell.number_format = 'yyyy-mm-dd hh:mm:ss'

    for column_idx, column_title in enumerate(fieldnames, 1):
        max_length = len(column_title)
        for row_idx in range(2, len(all_rows) + 2):
            cell = ws.cell(row=row_idx, column=column_idx)
            try:
                max_length = max(max_length, len(str(cell.value)))
            except Exception:
                continue
        ws.column_dimensions[get_column_letter(column_idx)].width = min(max_length + 2, 60)

    ws.freeze_panes = 'A2'

    wb.save(output_path)
    print(f"✓ Wrote combined workbook {output_path}")


def parse_order_datetime_value(raw_value: Any) -> Any:
    if isinstance(raw_value, datetime):
        return raw_value
    if raw_value in (None, ''):
        return None

    value = str(raw_value).strip()
    if not value:
        return None

    try:
        return datetime.fromisoformat(value.replace('Z', '+00:00'))
    except ValueError:
        pass

    for fmt in (
        '%Y-%m-%d %H:%M:%S',
        '%Y-%m-%d %H:%M',
        '%Y-%m-%d',
        '%m/%d/%Y %H:%M:%S',
        '%m/%d/%Y %H:%M',
        '%m/%d/%Y'
    ):
        try:
            return datetime.strptime(value, fmt)
        except ValueError:
            continue

    return raw_value


def parse_args():
    parser = argparse.ArgumentParser(
        description="Convert Staples order JSON files to TSV/XLSX."
    )
    parser.add_argument(
        '-p', '--profile',
        help="Optional profile name. When omitted, every subfolder in data/ is processed.",
        default=None
    )
    parser.add_argument(
        '-o', '--output-dir',
        help="Base directory for generated TSV/XLSX files. Defaults to each profile folder.",
        default=None
    )
    return parser.parse_args()


def main():
    args = parse_args()

    # Setup paths
    base_dir = Path(__file__).parent
    data_root = base_dir / 'data'

    if args.profile:
        targets = [(args.profile, data_root / args.profile)]
    else:
        targets = []
        if data_root.exists():
            if any(data_root.glob('*.json')):
                targets.append(('default', data_root))
            for subdir in sorted(p for p in data_root.iterdir() if p.is_dir()):
                if any(subdir.glob('*.json')):
                    targets.append((subdir.name, subdir))

    if not targets:
        print(f"Error: No profiles or JSON files found under {data_root}")
        sys.exit(1)

    output_base = Path(args.output_dir).expanduser() if args.output_dir else None
    multiple_profiles = len(targets) > 1
    combined_rows: List[Dict[str, Any]] = []

    for profile_name, profile_dir in targets:
        if not profile_dir.exists():
            print(f"Skipping '{profile_name}': missing directory {profile_dir}")
            continue

        json_files = sorted(profile_dir.glob('*.json'))
        if not json_files:
            print(f"Skipping '{profile_name}': no JSON files in {profile_dir}")
            continue

        if output_base:
            if multiple_profiles or not args.profile:
                output_dir = (output_base / profile_name)
            else:
                output_dir = output_base
        else:
            if args.profile:
                output_dir = profile_dir
            else:
                output_dir = profile_dir if profile_name != 'default' else base_dir

        output_dir.mkdir(parents=True, exist_ok=True)

        print(f"\nProcessing profile '{profile_name}' ({len(json_files)} file(s))...")

        all_rows = []

        for json_file in json_files:
            print(f"  - Reading {json_file.name}...")
            data = load_json_file(json_file)
            if not data:
                continue

            orders = data.get('orderDetailsList', []) or data.get('orders', [])
            if not orders:
                print(f"    No orders found in {json_file.name}")
                continue

            for order in orders:
                rows = parse_order_lines(order)
                all_rows.extend(rows)

            print(f"    Extracted {len(orders)} orders")

        if not all_rows:
            print(f"  ✗ No order data to export for '{profile_name}'")
            continue

        profile_rows_for_combined: List[Dict[str, Any]] = []
        for row in all_rows:
            row_copy = dict(row)
            row_copy['Profile'] = profile_name
            profile_rows_for_combined.append(row_copy)

        combined_rows.extend(profile_rows_for_combined)

        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        suffix = f"{profile_name}_{timestamp}" if profile_name != 'default' else timestamp

        print(f"  Generating output files ({len(all_rows)} rows)...")

        tsv_file = output_dir / f'orders_{suffix}.tsv'
        write_tsv(all_rows, tsv_file)

        if XLSX_AVAILABLE:
            xlsx_file = output_dir / f'orders_{suffix}.xlsx'
            write_xlsx(all_rows, xlsx_file)

    if XLSX_AVAILABLE and combined_rows:
        combined_dir = output_base if output_base else (args.profile and (base_dir / 'data' / args.profile) or base_dir / 'data')
        combined_dir.mkdir(parents=True, exist_ok=True)
        combined_path = combined_dir / f"orders_all_profiles_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        write_combined_sheet(combined_rows, combined_path)
    elif not XLSX_AVAILABLE:
        print("Skipping combined workbook (openpyxl not installed).")

    print("\n✅ Conversion complete!")


if __name__ == '__main__':
    main()
